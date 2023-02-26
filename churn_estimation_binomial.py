import math
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
from scipy.optimize import minimize, NonlinearConstraint
import openpyxl



#.......................My addded Code.........................................................................#



'''To add nonlinear constraints on the parameters r and c, you can use the scipy.optimize.minimize() 
function with the constraints parameter. The minimize() function allows you to optimize a function s
ubject to constraints on its parameters.'''

'''Here's an example of how you could modify your code to add nonlinear constraints on r and c: '''




#2. define a separate function to estimate the parameters c and r. (don't use sm.GLM() function in it. only use minimize() function)


def estimate_c_r(x_arr, y_arr):
    def neg_log_likelihood(params, x, y):
        c, r = params
        p = 1 - np.power(1 - c, x)
        likelihood = np.sum(y * np.log(p) + (1 - y) * np.log(1 - p))
        return -likelihood
    
   

def objective(params, x, y, cohort_customer_num):
    a, b, r, c = params
    p = np.power(c, x) * r
    likelihood = np.sum(binom.logpmf(y, p=p))
    return -1 * likelihood

def constraint(params):
        r, c = params
        C = 1 - math.pow(1 - c, x)
        return  r*C - c*R - rc 
    

    

def get_rate(x_arr, y_arr, cohort_customer_num):
    x = np.array(x_arr)
    y = np.array(y_arr)
    plt.scatter(x, y)
    plt.show()

   

  

    # set initial values for r and c
    r_init = 0.1
    c_init = 0.9
    # Set up initial parameter values and bounds
    params_init = [0.1, 0.1]
    bounds = ((0, 1), (0, None))
    # define the objective function and constraints
    params_init = [r_init, c_init]
    nonlinear_constraint = NonlinearConstraint(constraint, 0, np.inf)
    bounds = [(0, None), (1, None)]


   
    # optimize the objective function subject to the constraints
    res = minimize(objective, params_init, method='trust-constr', bounds=bounds, constraints=[nonlinear_constraint])

    nonlinear_constraint = NonlinearConstraint(constraint, 0, np.inf)
    bounds = [(None, None), (None, None), (0, None), (1, None)]
    # Extract the optimized parameter values
    c = res.x[0]
    r = res.x[1]

    # get the optimized parameters and return r and c
    a, b, r, c = res.x
    r_opt = r
    c_opt = c
    return r_opt, c_opt


'''In this code, the objective() function defines the objective function to be optimized,
 which computes the difference between the model predictions and the 
 target cohort_customer_num value. The constraint() function defines the nonlinear constraints on r and c.
 The get_rate() function sets the initial values for r and c, defines the bounds for all parameters, a
 nd uses the minimize() function with the trust-constr method to optimize the objective function subject to the constraints.'''

#..................................................My added Code............................................................#



#..........................................................#

'''To add two additional inputs n and R to the GLM_NonLinearConstrained function '''


def GLM_NonLinearConstrained(x_arr, y_arr, num, R):
    x = np.array(x_arr)
    y = np.array(y_arr)
    n = len(y_arr) # set n to length of y_arr
    plt.scatter(x, y)
    plt.show()
    
    # fit the model
    x = sm.tools.tools.add_constant(x)
    fit = sm.GLM(y, x, family=sm.families.Binomial(link=sm.families.links.log())).fit()

    # view the output of the model
    print(fit.params)
    a = math.exp(fit.params[0])
    b = math.exp(fit.params[1])
    print("-----",a,b)

    r = a / num
    c = 1 - b
    # C/c=[1-(1-c)^n]-1
    C = R * c / (1 - (1 - c)**n)
    return r, c, C


'''In this modified version of the function, we added two new inputs n and R to the function signature, 
and modified the last line of the function to calculate and return the value of C. 
The input n represents the value of the parameter n used in the calculation of C, 
and the input R represents the desired value of the ratio C/c.

Note that we assumed that the calculation of C is based on the following equation: C/c=[1-(1-c)^n]-1.'''

#............................................................................................................#

def get_month_repurchase_and_churn_rate(x_arr, y_arr, cohort_customer_num):
    x = np.array(x_arr)
    y = np.array(y_arr)
    plt.scatter(x, y)
    plt.show()

    # fit the model
    x = sm.tools.tools.add_constant(x)
    fit = sm.GLM(y, x, family=sm.families.Binomial(link=sm.families.links.log())).fit()

    # view the output of the model
    print(fit.params)
    a = math.exp(fit.params[0])
    b = math.exp(fit.params[1])
    print("-----",a,b)
    # a=同期群客户数*复购率
    #customerNum 同期群客户数
    r = a / cohort_customer_num

    # b=1-月流失率
    # 月流失率=1-b
    c = 1 - b
    return r, c

'''def p2f(x):
    return float(x.strip('%'))/100 '''
def p2f(x):
    if isinstance(x, str):
        return float(x.strip('%')) / 100
    elif isinstance(x, float):
        return x
    else:
        raise TypeError("Input must be a string or float")


def preprocess_data(data):
    M = float(data[3])
    r_true = p2f(data[1])
    c_true = p2f(data[0])
    y = np.array(data[4:], dtype=float)
    x = np.array(range(1, len(y)+1), dtype=float)
    not_nan = ~np.isnan(y)
    y = y[not_nan]
    x = x[not_nan]
    return M, r_true, c_true, x, y

def fit_data(df):
    N = len(df) - 1
    results = pd.DataFrame(columns=['cohort', 'cohort_customer_num', 'r_true', 'r_estimate', 'r_diff(%)', 'c_true', 'c_estimate', 'c_diff(%)'])
    for i in range(N):
        data = df.iloc[i+1, :]
        M, r_true, c_true, x, y = preprocess_data(data)
        if len(y) <= 2:
            print(f'Cohort_{i}: 数据点少于3个，无法估计')
        else:
            r, c = get_month_repurchase_and_churn_rate(x, y, M)
            r_diff = (r-r_true)/r_true*100
            c_diff = (c-c_true)/c_true*100
            print(f'Cohort_{i}: 估计值， 真值， 误差(%)')
            print(f'月复购率: {r:.4f}, {r_true}, {r_diff:.4f}')
            print(f'月流失率: {c:.4f}, {c_true}, {c_diff:.4f}')
            results.loc[len(results)] = [i, M, r_true, r, r_diff, c_true, c, c_diff]
    return results

if __name__ == '__main__':
    ## Toy example 测试
    x_arr = [1, 2, 3, 4, 5, 6, 7, 8]
    y_arr = [45.00, 40.50, 36.45, 30, 29.52, 26.57, 23.91, 21.52]
    num = 10
    R = sum(y_arr) / 500
    
    r, c, C = GLM_NonLinearConstrained(x_arr, y_arr, num, R)

    month = 9



    # 2、再计算 月复购率 月流失率
    x1=""
    
   # print(str(x1)+"月复购率",get_month_repurchase_rate(x_arr, y_arr, 100))
   
    print(str(x1)+"月流失率", GLM_NonLinearConstrained(x_arr, y_arr, num, R))


    s = GLM_NonLinearConstrained(x_arr, y_arr, num, R)

    print("类型",type(s))
    for value in s:
        if math.isnan(value):
           print(f"{value} is NaN")
   # print(math.isnan(s))

    # 不满一年计算方式
    # 1、先计算a b的值 预测下个月复购数
    # 2、再计算月流失率 月复购率
    # 3、再计算年流失率 年复购率
    
    
    print(1 - (1 - 0.2) **(1/ 12))
    


# Load the Excel file
workbook = openpyxl.load_workbook('churn_data_2.xlsx')
# Select the source and destination sheets
source_sheet = workbook['Number from Binomial']
destination_sheet = workbook['Number from Binomial (fixed ran']
target_sheet = workbook['Estimate for Sheet4']


# estimate the c and r for each row in sheet Number from Binomial (fixed ran
# Get the x and y values from the destination sheet


x_arr = []
y_arr = []
for column in range(2, 24):
    value = destination_sheet.cell(row=1, column=column).value
    if value is not None:
        if isinstance(value, str):
            value = value.replace(',', '')
        y_arr.append(int(value))
        x_arr.append(column - 1)
        
        
        
        
        

#4. compare the estimated c and r with the real c and r and output to the last sheet using the traditional GLM package.
# Get the real values of c and r from the target sheet

real_c = target_sheet.cell(row=23, column=2).value
real_r = target_sheet.cell(row=23, column=3).value
#5. compare the estimated c and r with the real c and r and output to the last sheet using 
#our new code (regression with nonlinear constraint)
# Output the comparison to the last sheet
output_sheet = workbook.create_sheet('Comparison')
output_sheet.cell(row=1, column=1).value = 'Parameter'
output_sheet.cell(row=1, column=2).value = 'Estimated'
output_sheet.cell(row=1, column=3).value = 'Real'
output_sheet.cell(row=2, column=1).value = 'c'
output_sheet.cell(row=2, column=3).value = real_c
output_sheet.cell(row=3, column=1).value = 'r'

# Copy the values from the source sheet to the destination sheet
for column in range(1, 24):
    source_value = source_sheet.cell(row=1, column=column).value
    destination_sheet.cell(row=1, column=column).value = source_value

# Change the values in the target sheet's column
for row in range(2, 23):
    target_sheet.cell(row=row, column=8).value = row * 10  # Change this to your desired calculation or rule
# Save the updated workbook
workbook.save('ata.xlsx')


